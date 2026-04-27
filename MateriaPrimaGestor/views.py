from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import get_template
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout
from django.contrib import messages
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db import transaction
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from reportlab.pdfgen import canvas
from copy import copy
from .forms import *
from .models import *
import os

def inicio(request):
    return render(request, "inicio.html")


# Registro
def register_view(request):
    if request.method == "POST":
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, "Usuario creado correctamente")
            return redirect("inicio")
        else:
            messages.error(request, "Corrige los errores")
    else:
        form = UserCreationForm()

    return render(request, "register.html", {"form": form})
# Login
def login_view(request):
    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()  # obtiene al usuario autenticado
            login(request, user)
            messages.success(request, f"Bienvenido {user.username}")
            return redirect("inicio")
        else:
            messages.error(request, "Credenciales inválidas")
    else:
        form = AuthenticationForm()
    
    return render(request, "login.html", {"form": form})

# Logout
def logout_view(request):
    logout(request)
    messages.info(request, "Sesión cerrada correctamente")
    return redirect("inicio")

############## CRUD de proveedor
@login_required
def listar_proveedores(request):
    proveedores = Proveedor.objects.all().order_by("codigo")
    return render(request, "proveedor/listar.html", {"proveedores": proveedores})

@login_required
def crear_proveedor(request):
    form = ProveedorForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, "Proveedor creado correctamente.")
        return redirect("listar_proveedores")
    return render(request, "proveedor/form.html", {"form": form, "titulo": "Crear Proveedor"})

@login_required
def editar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    form = ProveedorForm(request.POST or None, instance=proveedor)
    if form.is_valid():
        form.save()
        messages.success(request, "Proveedor actualizado correctamente.")
        return redirect("listar_proveedores")
    return render(request, "proveedor/form.html", {"form": form, "titulo": "Editar Proveedor"})

@login_required
def eliminar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == "POST":
        proveedor.delete()
        messages.success(request, "Proveedor eliminado correctamente.")
        return redirect("listar_proveedores")
    return render(request, "proveedor/eliminar.html", {"proveedor": proveedor})

############## CRUD de MateriaPrima
@login_required
def listar_materias(request):
    materias = MateriaPrima.objects.all().order_by("codigo")
    return render(request, "materia/listar.html", {"materias": materias})

@login_required
def materias_pdf(request):
    materias = MateriaPrima.objects.all().order_by("codigo")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=inventario_materias.pdf"

    p = canvas.Canvas(response)

    y = 800
    p.drawString(200, y, "REPORTE DE MATERIAS PRIMAS")
    y -= 40

    for m in materias:
        texto = f"{m.codigo} - {m.nombre} - Stock: {m.cantidad}"
        p.drawString(50, y, texto)
        y -= 20

        if y < 50:
            p.showPage()
            y = 800

    p.save()
    return response

@login_required
def crear_materia(request):
    form = MateriaPrimaCreateForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, "Materia prima creada.")
        return redirect("listar_materias")
    return render(request, "materia/form.html", {"form": form, "titulo": "Nueva Materia Prima"})

@login_required
def editar_materia(request, pk):
    materia = get_object_or_404(MateriaPrima, pk=pk)
    form = MateriaPrimaUpdateForm(request.POST or None, instance=materia)
    if form.is_valid():
        form.save()
        messages.success(request, "Materia prima actualizada.")
        return redirect("listar_materias")
    return render(request, "materia/form.html", {"form": form, "titulo": "Editar Materia Prima"})

@login_required
def ajustar_materia(request, pk):
    materia = get_object_or_404(MateriaPrima, pk=pk)
    form = MateriaPrimaAjusteForm(request.POST or None)
    if form.is_valid():
        cantidad = form.cleaned_data["cantidad"]
        tipo = form.cleaned_data["tipo"]
        if tipo == "SUMAR":
            materia.cantidad += cantidad
        else:
            if cantidad > materia.cantidad:
                messages.error(request, "No puedes dejar stock negativo.")
                return redirect("ajustar_materia", pk=pk)
            materia.cantidad -= cantidad
        materia.save()
        messages.success(request, "Stock actualizado correctamente.")
        return redirect("listar_materias")
    return render(request, "materia/ajustar.html", {"form": form, "materia": materia})

@login_required
def eliminar_materia(request, pk):
    materia = get_object_or_404(MateriaPrima, pk=pk)
    if request.method == "POST":
        materia.delete()
        messages.success(request, "Materia prima eliminada.")
        return redirect("listar_materias")
    return render(request, "materia/eliminar.html", {"materia": materia})

############## CRUD de Compra 
@login_required
def listar_compras(request):
    compras = Compra.objects.select_related("proveedor", "materia_prima").all().order_by("-codigo")
    return render(request, "compra/listar.html", {"compras": compras})

@login_required
@transaction.atomic
def crear_compra(request):
    form = CompraForm(request.POST or None)
    if form.is_valid():
        compra = form.save(commit=False)
        materia = compra.materia_prima
        # Aumentar inventario
        materia.cantidad += compra.cantidad
        # Actualizar costo al último comprado
        materia.costo = (Decimal(compra.costo) / Decimal(compra.cantidad)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        materia.save()
        compra.save()
        messages.success(request, "Compra registrada correctamente.")
        return redirect("listar_compras")

    return render(request, "compra/form.html", {"form": form, "titulo": "Registrar Compra"})

@login_required
@transaction.atomic
def editar_compra(request, pk):
    compra = get_object_or_404(Compra, pk=pk)
    materia = compra.materia_prima
    cantidad_anterior = compra.cantidad
    form = CompraForm(request.POST or None, instance=compra)

    if form.is_valid():
        compra_editada = form.save(commit=False)
        # Quitar efecto anterior
        materia.cantidad -= cantidad_anterior
        # Aplicar nuevo efecto
        materia.cantidad += compra_editada.cantidad
        # Actualizar costo
        materia.costo = (Decimal(compra_editada.costo) / Decimal(compra_editada.cantidad)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        materia.save()
        compra_editada.save()
        messages.success(request, "Compra actualizada correctamente.")
        return redirect("listar_compras")
    return render(request, "compra/form.html", {"form": form, "titulo": "Editar Compra"})

@login_required
def eliminar_compra(request, pk):
    compra = get_object_or_404(Compra, pk=pk)
    if request.method == "POST":
        compra.delete()
        messages.success(request, "Compra eliminada.")
        return redirect("listar_compras")
    return render(request, "compra/eliminar.html", {"compra": compra})

############## CRUD de producto terminado 
@login_required
def listar_productos(request):
    productos = ProductoTerminado.objects.all().order_by("codigo")
    return render(request, "producto/listar.html", {"productos": productos})

@login_required
def productos_pdf(request):
    productos = ProductoTerminado.objects.all().order_by("codigo")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=inventario_productos.pdf"

    p = canvas.Canvas(response)

    y = 800
    p.drawString(200, y, "REPORTE DE PRODUCTOS")
    y -= 40

    for pdt in productos:
        texto = f"{pdt.codigo} - {pdt.nombre} - Stock: {pdt.cantidad}"
        p.drawString(50, y, texto)
        y -= 20

        if y < 50:
            p.showPage()
            y = 800

    p.save()
    return response

@login_required
def crear_producto(request):
    form = ProductoForm(request.POST or None)
    formset = DetalleProductoFormSet(request.POST or None)
    if request.method == "POST":
        if form.is_valid() and formset.is_valid():
            producto = form.save()
            formset.instance = producto
            formset.save()
            messages.success(request, "Producto y receta creados correctamente.")
            return redirect("listar_productos")
        else:
            messages.error(request, "Corrige los errores del formulario.")
    return render(request, "producto/form.html", {
        "form": form,
        "formset": formset,
        "titulo": "Crear Producto",
    })


@login_required
def editar_producto(request, pk):
    producto = get_object_or_404(ProductoTerminado, pk=pk)
    form = ProductoForm(request.POST or None, instance=producto)
    formset = DetalleProductoFormSet(request.POST or None, instance=producto)
    if request.method == "POST":
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, "Receta actualizada correctamente.")
            return redirect("listar_productos")
        else:
            messages.error(request, "Corrige los errores del formulario.")
    return render(request, "producto/form.html", {
        "form": form,
        "formset": formset,
        "titulo": "Editar Producto",
    })

@login_required
def eliminar_producto(request, pk):
    producto = get_object_or_404(ProductoTerminado, pk=pk)

    if request.method == "POST":
        producto.delete()
        messages.success(request, "Producto eliminado.")
        return redirect("listar_productos")

    return render(request, "producto/eliminar.html", {"producto": producto})

############## CRUD de produccion
@login_required
def listar_producciones(request):
    producciones = ProduccionProducto.objects.select_related("producto").all().order_by("-codigo")
    return render(request, "produccion/listar.html", {"producciones": producciones})

@login_required
def listar_producciones(request):
    producciones = ProduccionProducto.objects.select_related("producto").all().order_by("-codigo")
    return render(request, "produccion/listar.html", {"producciones": producciones})

@login_required
@transaction.atomic
def crear_produccion(request):
    form = ProduccionForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        producto = form.cleaned_data["producto"]
        cantidad_producir = form.cleaned_data["cantidad"]
        cantidad_unidad = form.cleaned_data["cantidad_unidad"]
        receta = DetalleProducto.objects.filter(producto=producto)

        if not receta.exists():
            messages.error(request, "El producto no tiene receta.")
            return render(request, "produccion/form.html", {"form": form, "titulo": "Nueva Producción"})

        # VALIDAR STOCK
        for item in receta:
            requerido = item.cantidad * cantidad_producir
            if item.materia_prima.cantidad < requerido:
                messages.error(request, f"Stock insuficiente de {item.materia_prima.nombre}")
                return render(request, "produccion/form.html", {"form": form, "titulo": "Nueva Producción"})

        # GUARDAR
        produccion = form.save()
        producto.cantidad += cantidad_unidad
        producto.save()

        costo_total = Decimal("0")
        for item in receta:
            materia = item.materia_prima
            requerido = item.cantidad * cantidad_producir
            materia.cantidad -= requerido
            materia.save()
            costo = materia.costo
            costo_total += requerido * costo
            ProduccionDetalle.objects.create(
                produccion=produccion,
                materia_prima=materia,
                cantidad=requerido,
                costo=costo
            )

        messages.success(request, f"Producción registrada. Costo total: {round(costo_total, 2)}")
        return redirect("listar_producciones")

    return render(request, "produccion/form.html", {"form": form, "titulo": "Nueva Producción"})


@login_required
@transaction.atomic
def editar_produccion(request, pk):
    produccion = get_object_or_404(ProduccionProducto, pk=pk)
    form = ProduccionForm(request.POST or None, instance=produccion)

    if request.method == "POST" and form.is_valid():
        producto = form.cleaned_data["producto"]
        nueva_cantidad = form.cleaned_data["cantidad"]
        cantidad_unidad = form.cleaned_data["cantidad_unidad"]
        receta = DetalleProducto.objects.filter(producto=producto)

        if not receta.exists():
            messages.error(request, "El producto no tiene receta.")
            return render(request, "produccion/form.html", {"form": form, "titulo": "Editar Producción"})

        # 1. VALIDAR STOCK SIMULANDO DEVOLUCIÓN EN MEMORIA
        detalles_anteriores = ProduccionDetalle.objects.filter(produccion=produccion)
        stock_simulado = {
            det.materia_prima.pk: det.materia_prima.cantidad + det.cantidad
            for det in detalles_anteriores
        }
        for item in receta:
            mp_pk = item.materia_prima.pk
            stock_disponible = stock_simulado.get(mp_pk, item.materia_prima.cantidad)
            if stock_disponible < item.cantidad * nueva_cantidad:
                messages.error(request, f"Stock insuficiente de {item.materia_prima.nombre}")
                return render(request, "produccion/form.html", {"form": form, "titulo": "Editar Producción"})

        # 2. DEVOLVER STOCK ANTERIOR
        for det in detalles_anteriores:
            det.materia_prima.cantidad += det.cantidad
            det.materia_prima.save()

        produccion.producto.cantidad -= produccion.cantidad_unidad
        produccion.producto.save()
        detalles_anteriores.delete()

        # 3. GUARDAR Y APLICAR NUEVA PRODUCCIÓN
        produccion = form.save()
        produccion.producto.cantidad += cantidad_unidad
        produccion.producto.save()

        for item in receta:
            materia = item.materia_prima
            materia.refresh_from_db()
            requerido = item.cantidad * nueva_cantidad
            materia.cantidad -= requerido
            materia.save()
            ProduccionDetalle.objects.create(
                produccion=produccion,
                materia_prima=materia,
                cantidad=requerido,
                costo=materia.costo
            )

        messages.success(request, "Producción recalculada correctamente.")
        return redirect("listar_producciones")

    return render(request, "produccion/form.html", {"form": form, "titulo": "Editar Producción"})

@login_required
def eliminar_produccion(request, pk):
    produccion = get_object_or_404(ProduccionProducto, pk=pk)
    if request.method == "POST":
        produccion.delete()
        messages.success(request, "Producción eliminada.")
        return redirect("listar_producciones")
    return render(request, "produccion/eliminar.html", {"produccion": produccion})

@login_required
def ajustar_producto(request, pk):
    producto = get_object_or_404(ProductoTerminado, pk=pk)
    form = ProductoAjusteForm(request.POST or None)
    if form.is_valid():
        cantidad = form.cleaned_data["cantidad"]
        tipo = form.cleaned_data["tipo"]
        if tipo == "SUMAR":
            producto.cantidad += cantidad
        else:
            if cantidad > producto.cantidad:
                messages.error(request, "No puedes dejar stock negativo.")
                return redirect("ajustar_producto", pk=pk)
            producto.cantidad -= cantidad
        producto.save()
        messages.success(request, "Stock actualizado correctamente.")
        return redirect("listar_productos")
    return render(request, "producto/ajustar.html", {"form": form, "producto": producto})

@login_required
def exportar_produccion_excel(request, pk):
    # Datos
    produccion = get_object_or_404(ProduccionProducto, pk=pk)
    detalles = list(ProduccionDetalle.objects.filter(produccion=produccion))
    codigo = f'PR-PRO-{produccion.codigo:03d}'

    # Cargar plantilla
    ruta = os.path.join(settings.BASE_DIR, 'templates', 'PLANTILLA.xlsx')
    wb = load_workbook(ruta)
    ws = wb["PRODUCCION"]

    # === METADATOS ===
    ws["H1"] = codigo
    ws["H2"] = "1"
    ws["H3"] = datetime.now().strftime("%d/%m/%Y")
    ws["H4"] = "1 de 1"
    ws["H5"] = produccion.fecha_creacion.strftime("%d/%m/%Y")
    ws["E5"] = produccion.producto.codigo
    ws["C6"] = produccion.producto.nombre
    ws["C7"] = ""
    ws["E7"] = produccion.cantidad

    # === PARTE DINÁMICA ===
    fila_inicio = 9
    cantidad = len(detalles)

    if cantidad > 1:
        cantidad_a_insertar = cantidad - 1

        # 1. Desplazar merges que estén debajo (fila 10 al 30 y más abajo)
        for merged in list(ws.merged_cells.ranges):
            if merged.min_row >= fila_inicio + 1:
                merged.shift(row_shift=cantidad_a_insertar)

        # 2. Insertar filas vacías
        ws.insert_rows(fila_inicio + 1, amount=cantidad_a_insertar)

        # 3. Copiar SOLO formato (incluyendo dentro de merges como E:F)
        for offset in range(1, cantidad):
            fila_nueva = fila_inicio + offset

            # Altura de fila
            ws.row_dimensions[fila_nueva].height = ws.row_dimensions[fila_inicio].height

            # Copiar estilos celda por celda (CORREGIDO para merges)
            for col in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=fila_inicio, column=col)
                dst_cell = ws.cell(row=fila_nueva, column=col)

                dst_cell.value = None

                # Si es celda combinada, tomamos el estilo real del top-left
                if isinstance(src_cell, MergedCell):
                    for m in list(ws.merged_cells.ranges):
                        if (m.min_row == fila_inicio and 
                            m.min_col <= col <= m.max_col):
                            src_cell = ws.cell(row=m.min_row, column=m.min_col)
                            break

                if src_cell.has_style:
                    dst_cell.font = copy(src_cell.font)
                    dst_cell.border = copy(src_cell.border)
                    dst_cell.fill = copy(src_cell.fill)
                    dst_cell.number_format = src_cell.number_format
                    dst_cell.protection = copy(src_cell.protection)
                    dst_cell.alignment = copy(src_cell.alignment)

    # 4. Replicar merges horizontales (E:F y cualquier otro) SOLO en las filas nuevas
    for offset in range(1, cantidad):
        fila = fila_inicio + offset
        for merged in list(ws.merged_cells.ranges):
            if merged.min_row == fila_inicio and merged.max_row == fila_inicio:
                start = get_column_letter(merged.min_col)
                end = get_column_letter(merged.max_col)
                ws.merge_cells(f"{start}{fila}:{end}{fila}")

    # 5. Rellenar datos reales
    for idx, det in enumerate(detalles, start=fila_inicio):
        fila = idx
        ws[f"A{fila}"] = idx - fila_inicio + 1
        ws[f"B{fila}"] = det.materia_prima.codigo
        ws[f"C{fila}"] = det.materia_prima.nombre
        ws[f"D{fila}"] = det.cantidad
        valor = (det.cantidad * produccion.cantidad).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
        ws[f"E{fila}"] = float(valor)
        ws[f"G{fila}"] = float(valor*1000)

    # === DESCARGA ===
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Produccion_{codigo}.xlsx"'
    wb.save(response)
    return response

# APIS json
def get_receta(request, producto_id):
    receta = DetalleProducto.objects.filter(producto_id=producto_id).values(
        'materia_prima__nombre',
        'materia_prima__unidad',
        'materia_prima__cantidad',  # stock actual
        'materia_prima__costo',
        'cantidad'  # cantidad por unidad de producto
    )
    return JsonResponse(list(receta), safe=False)

def get_materias_primas(request):
    materias = MateriaPrima.objects.all().order_by('nombre').values(
        'id', 'nombre', 'unidad', 'cantidad', 'costo'
    )
    return JsonResponse(list(materias), safe=False)